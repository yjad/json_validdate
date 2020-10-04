import json
from openpyxl import Workbook
file_name = r'.\response.json'

def print_keys(prefix_id, parent_dict_key, dictx):

    for p_id, p_info in dictx.items():
        v = "-> " + p_info if type(p_info) == str else ""
        print(prefix_id, "-", parent_dict_key, ":", p_id, v)

        if type(p_info) == list:
            for biller in p_info:   # array
                print_keys(prefix_id+1, p_id, biller)
        elif type(p_info) == dict:
            print_keys(prefix_id + 1, p_id, p_info)
    if prefix_id == 2:
        print("-----", 20 * str(prefix_id), "------")

def print_values(prefix, dictx):

    for p_id, p_info in dictx.items():
        print (p_id, type(p_info))
        if type(p_info) == dict:
            for k, v in p_info.items():
                print(prefix, k, v)

        elif type(p_info) == list:
            for biller in p_info:  # array
                print_values(p_id, biller)
                # for key in biller:
                #     print(p_id, key + ':', biller[key])
        else:
            print(prefix, p_id, p_info)



def list_billers (dicx):

    headers = dicx[0]
    arr=[]
    for k, v in headers.items():
        if type(v) not in (dict, list) :
            arr.append(k)
    print (arr)

    for biller in dicx:
        arr=[]
        for k, v in biller.items():
            if type(v) not in (dict, list) :
                arr.append(v)
        print (arr)


def list_biller_info (dicx, key):

    headers = dicx[0][key]
    if type(headers) == list:
        # arr=[]
        # for k, v in headers[0].items():
        #     if type(v) not in (dict, list) :
        #         arr.append(k)
        # print (arr)`

        for item in dicx:
            for biller in item[key]:
                print(biller.keys())
                arr=[]
                for k, v in biller.items():
                    if type(v) not in (dict, list) :
                        arr.append(v)
                print(arr)


def load_json_file():
    with open(file_name, 'r', encoding ='UTF-8') as myfile:
        data=myfile.read()
    myfile.close
    # parse file
    billers = json.loads(data)

    return billers


def format_dict_file(dictx):
    fp = open('biller.json', 'wt')
    json.dump(dictx, fp, indent=4)


def print_fawry_billers():
    billers = load_json_file()
    print_keys(0, "billers", billers)


def fawry_billers_excel():
    billers = load_json_file()
    wb = Workbook()
    ws = wb.active

    for biller_rec in billers['biller_records']:
        for biller_info in biller_rec['biller_information']:
            row = []
            header = []
            row.append(biller_rec['id'])
            row.append(biller_rec['name'])
            header.append('id')
            header.append('name')     
            dict_to_row(biller_info, row, header) 
            #print (type(header), biller_rec['id'])
            if header:
                ws.append(header)
                ws.append(row)
    wb.save(file_name +".xlsx")
        
    return

def not_used():
    billers = load_json_file()
    wb = Workbook()
    ws = wb.active
    row, header = dict_to_row(billers['biller_records'][0])
    ws.append(header)
    ws.append(row)
    for b_rec in billers['biller_records']:
        row,header = dict_to_row(b_rec)
        ws.append(header)
        ws.append(row)

    sheet1 = wb.create_sheet()
    row, header = dict_to_row(billers['biller_records'][0]['biller_information'][0])
    sheet1.append(header)
    sheet1.append(row)
    for b_rec in billers['biller_records']:
        for b in b_rec['biller_information']:
            row, header = dict_to_row(b)
            sheet1.append(header)
            sheet1.append(row)

    sheet2 = wb.create_sheet()
    sheet2 = sheet2.active
    #for b_rec in billers['biller_records']:
    #    row, header = dict_to_row(b_rec)
    #    sheet2.append(header)
    #    sheet2.append(row)

    
    wb.save(file_name +".xlsx")

    
def list_biller_records(biller_records):

    row = []
    header = []
    for brec in biller_records:   # list
        for brec_id, brec_value in b_rec.items():
            if type(brec_value) == str:
                header.append(brec_id)
                row.append(brec_value)
                continue
            for binfo in brec_value.get('biller_information'): # list of biller information
                for binfo_id, binfo_value in binfo.items():   # dict items
                    if type(binfo_value) == str:
                        header.append("biller_information: " + binfo_value)
                        row.append(binfo_id)
                    elif binfo_id == 'payment_rules': #dict
                        for prule_id, prule_value in binfo_value:
                            header.append("biller_information: payment_rules: " + prule_info)
                            row.append(prule_id)
                    elif binfo_id == 'fees': #dict
                        for fee in binfo_value:# list
                            for fee_id, fee_value in fee:
                                if fee_id == 'tiers': # list
                                    for tier in fee_value:
                                        for tier_id, tier_value in tier.items():
                                            header.append("tiers: " + tier_id)
                                            row.append(tier_value)
                                else:
                                    header.append("fees: " + fee_id_id)
                                    row.append(fee_value)
                        
                row, header = dict_to_row(b)
                sheet1.append(header)
                sheet1.append(row)

    
def dict_to_row(dictx, row, header): # dict or list of dict

    if type(dictx) == dict:
        for p_id, p_info in dictx.items():
            if p_info == None: 
                continue
            if type(p_info) == str:
                header.append(p_id)
                row.append(p_info)
            elif type(p_info) in (list, dict):
                dict_to_row(p_info, row, header)

    elif type(dictx) == list:  # dict in a dict
        for list_item in dictx:
            if list_item == None:
                continue
            dict_to_row(list_item, row, header)
    return

def get_biller_record(biller_name):
    billers = load_json_file()
    for biller_rec in billers['biller_records']:
        if biller_rec.get('name') == biller_name:
            for biller_info in biller_rec['biller_information']:
                row = []
                header = []
                row.append(biller_rec['id'])
                row.append(biller_rec['name'])
                header.append('id')
                header.append('name')
                dict_to_row(biller_info, row, header)
                #print (type(header), biller_rec['id'])
                if header:
                    print(header)
                    print(row)
            return
    print (f"Biller rec {biller_name} not found")